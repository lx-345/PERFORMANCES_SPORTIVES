from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series

# Importation des outils de design depuis notre module local
from .styles import (
    ajouter_filtre_deroulant, 
    appliquer_titre_principal, 
    appliquer_texte_intro, 
    appliquer_carte_kpi, 
    styliser_tableau_donnees,
    configurer_graphique_pro
)

def injecter_kpis_filtres_analyse(ws, max_data_row):
    """Génère les 3 cartes KPIs en haut de la page Analyse"""
    max_dash_row = 44 + max_data_row - 1
    
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=3)
    ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=3)
    appliquer_carte_kpi(ws, 7, 2, "VMA MOYENNE (km/h)", f'=IFERROR(AVERAGE(H45:H{max_dash_row}), 0)', "0.0")

    ws.merge_cells(start_row=7, start_column=5, end_row=7, end_column=6)
    ws.merge_cells(start_row=8, start_column=5, end_row=8, end_column=6)
    appliquer_carte_kpi(ws, 7, 5, "MEILLEUR TEMPS", f'=IFERROR(MIN(G45:G{max_dash_row}) / 1440, 0)')
    ws.cell(row=8, column=5).number_format = 'hh:mm:ss'

    ws.merge_cells(start_row=7, start_column=8, end_row=7, end_column=9)
    ws.merge_cells(start_row=8, start_column=8, end_row=8, end_column=9)
    appliquer_carte_kpi(ws, 7, 8, "VOLUME TOTAL (km)", f'=SUM(F45:F{max_dash_row})', "#,##0.0")

def integrer_graphiques_analyse(wb, ws_dash, max_pivot_row, max_data_row):
    """Assemble et place les graphiques de performances"""
    ws_pivot, ws_data = wb["Data_Pivot"], wb["DATA"]
    cats_vol = Reference(ws_pivot, min_col=1, min_row=2, max_row=max_pivot_row)

    # 1. Volume
    chart_vol = configurer_graphique_pro(BarChart())
    chart_vol.type = "col"
    chart_vol.title = "Volume d'Entraînement par Trimestre (km)"
    chart_vol.height, chart_vol.width = 7.5, 13
    chart_vol.add_data(Reference(ws_pivot, min_col=2, min_row=1, max_row=max_pivot_row), titles_from_data=True)
    chart_vol.set_categories(cats_vol)
    if chart_vol.series: chart_vol.series[0].graphicalProperties.solidFill = "E16B1A"
    ws_dash.add_chart(chart_vol, "B11")

    # 2. VMA
    chart_vma = configurer_graphique_pro(BarChart())
    chart_vma.type = "col"
    chart_vma.title = "Évolution de la VMA Moyenne par Trimestre"
    chart_vma.height, chart_vma.width = 7.5, 13
    chart_vma.add_data(Reference(ws_pivot, min_col=3, min_row=1, max_row=max_pivot_row), titles_from_data=True)
    chart_vma.set_categories(cats_vol)
    if chart_vma.series: chart_vma.series[0].graphicalProperties.solidFill = "F39C12"
    ws_dash.add_chart(chart_vma, "J11")

    # 3. Polarisation
    chart_scatter = configurer_graphique_pro(ScatterChart())
    chart_scatter.title = "Matrice Polarisation : Indice Effort vs Stress"
    chart_scatter.height, chart_scatter.width = 7.5, 13
    xvalues_polar = Reference(ws_data, min_col=9, min_row=2, max_row=max_data_row)
    for nom, col, couleur in [("Intense", 12, "C0392B"), ("Modéré", 13, "E16B1A"), ("Faible", 14, "F8C471")]:
        serie = Series(Reference(ws_data, min_col=col, min_row=2, max_row=max_data_row), xvalues_polar, title=nom)
        serie.marker.symbol, serie.marker.graphicalProperties.solidFill = "circle", couleur
        serie.graphicalProperties.line.noFill = True
        chart_scatter.series.append(serie)
    ws_dash.add_chart(chart_scatter, "B27")

    # 4. Double Axe (Lissé)
    chart_double_y = configurer_graphique_pro(LineChart())
    chart_double_y.title = "Efficience Chronologique : Temps vs Indice d'Effort"
    chart_double_y.height, chart_double_y.width = 7.5, 13
    chart_double_y.y_axis.title = "Temps (min)"
    
    chart_double_y.add_data(Reference(ws_data, min_col=15, min_row=1, max_row=max_data_row), titles_from_data=True)
    chart_double_y.set_categories(Reference(ws_data, min_col=2, min_row=2, max_row=max_data_row))
    if chart_double_y.series: 
        chart_double_y.series[0].graphicalProperties.line.solidFill = "2980B9"
        chart_double_y.series[0].smooth = True 
    
    chart_effort_sec = LineChart()
    chart_effort_sec.add_data(Reference(ws_data, min_col=16, min_row=1, max_row=max_data_row), titles_from_data=True)
    chart_effort_sec.y_axis.axId = 200
    chart_effort_sec.y_axis.crosses = "max"
    chart_effort_sec.y_axis.title = "Indice Effort"
    if chart_effort_sec.series:
        chart_effort_sec.series[0].graphicalProperties.line.solidFill = "C0392B"
        chart_effort_sec.series[0].smooth = True 
    
    chart_double_y += chart_effort_sec
    ws_dash.add_chart(chart_double_y, "J27")

def construire_page_analyse(wb, max_data_row, len_y, len_t, len_d, max_pivot_row):
    """Construit la page complète Analyse"""
    ws_dash = wb.create_sheet(title="Analyse")
    
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for row in ws_dash.iter_rows(min_row=1, max_row=250, min_col=1, max_col=22):
        for cell in row: cell.fill = fill_white
    ws_dash.sheet_view.showGridLines = False

    ajouter_filtre_deroulant(ws_dash, 2, 2, "Année", f"=RESSOURCES!$A$1:$A${len_y}", "Tous")
    ajouter_filtre_deroulant(ws_dash, 2, 5, "Trimestre", f"=RESSOURCES!$B$1:$B${len_t}", "Tous")
    ajouter_filtre_deroulant(ws_dash, 2, 8, "Distance Cible", f"=RESSOURCES!$C$1:$C${len_d}", "Tous")

    appliquer_titre_principal(ws_dash.cell(row=4, column=2), "Analyse du profil de l’athlète")
    ws_dash.merge_cells(start_row=5, start_column=2, end_row=5, end_column=10)
    appliquer_texte_intro(ws_dash.cell(row=5, column=2), "Sélectionnez vos critères pour actualiser les graphiques.")

    start_row_table = 44
    headers = ["Date", "Année", "Trimestre", "Distance Cible", "Distance (km)", "Temps (min)", "VMA (km/h)"]
    for col_idx, text in enumerate(headers, start=2): ws_dash.cell(row=start_row_table, column=col_idx, value=text)
        
    current_row = start_row_table + 1
    for i in range(2, max_data_row + 1):
        cond = f'AND(OR(Analyse!$B$2="Tous", DATA!C{i}&""=Analyse!$B$2&""), OR(Analyse!$E$2="Tous", DATA!D{i}&""=Analyse!$E$2&""), OR(Analyse!$H$2="Tous", DATA!E{i}&""=Analyse!$H$2&""))'
        
        ws_dash.cell(row=current_row, column=2, value=f'=IF({cond}, DATA!B{i}, "")') 
        ws_dash.cell(row=current_row, column=3, value=f'=IF({cond}, DATA!C{i}, "")') 
        ws_dash.cell(row=current_row, column=4, value=f'=IF({cond}, DATA!D{i}, "")') 
        ws_dash.cell(row=current_row, column=5, value=f'=IF({cond}, DATA!E{i}, "")') 
        ws_dash.cell(row=current_row, column=6, value=f'=IF({cond}, DATA!F{i}, "")') 
        ws_dash.cell(row=current_row, column=7, value=f'=IF({cond}, DATA!G{i}, "")') 
        ws_dash.cell(row=current_row, column=8, value=f'=IF({cond}, DATA!H{i}, "")') 
        
        ws_dash.cell(row=current_row, column=2).number_format = 'YYYY-MM-DD'
        ws_dash.cell(row=current_row, column=6).number_format = '0.0'
        ws_dash.cell(row=current_row, column=7).number_format = '0.0'
        ws_dash.cell(row=current_row, column=8).number_format = '0.0'
        current_row += 1

    styliser_tableau_donnees(ws_dash, start_row_table, current_row - 1, start_col=2, end_col=8)
    
    # Appel des KPIs et Graphiques
    injecter_kpis_filtres_analyse(ws_dash, max_data_row)
    integrer_graphiques_analyse(wb, ws_dash, max_pivot_row, max_data_row)
    
    ws_dash.column_dimensions['A'].width = 3
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']: ws_dash.column_dimensions[col].width = 15