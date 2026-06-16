from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

# Importation des outils de design depuis notre module local
from .styles import (
    ajouter_filtre_deroulant, 
    appliquer_titre_principal, 
    appliquer_texte_intro, 
    appliquer_carte_kpi, 
    styliser_tableau_donnees
)

def configurer_graphique_pro(chart):
    """Applique le style professionnel standard aux graphiques"""
    chart.style = 14 
    if chart.legend:
        chart.legend.position = 'b' 
    return chart

def injecter_kpis_charge(ws, max_data_row):
    """Génère les 3 cartes KPIs en interrogeant directement l'onglet DATA via les filtres de la feuille"""
    
    # Préparation des conditions dynamiques d'écoute des menus déroulants
    cond_y = 'IF($B$2="Tous", "*", $B$2)'
    cond_t = 'IF($E$2="Tous", "*", $E$2)'
    cond_d = 'IF($H$2="Tous", "*", $H$2)'

    # KPI 1 : RATIO ACWR MOYEN (Interroge la colonne U de DATA)
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=3)
    ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=3)
    formule_acwr = f'=IFERROR(AVERAGEIFS(DATA!$U$2:$U${max_data_row}, DATA!$C$2:$C${max_data_row}, {cond_y}, DATA!$D$2:$D${max_data_row}, {cond_t}, DATA!$E$2:$E${max_data_row}, {cond_d}), 1.00)'
    appliquer_carte_kpi(ws, 7, 2, "RATIO ACWR MOYEN", formule_acwr, "0.00")

    # KPI 2 : STRESS MÉCANIQUE MOYEN (Interroge la colonne J de DATA)
    ws.merge_cells(start_row=7, start_column=5, end_row=7, end_column=6)
    ws.merge_cells(start_row=8, start_column=5, end_row=8, end_column=6)
    formule_stress = f'=IFERROR(AVERAGEIFS(DATA!$J$2:$J${max_data_row}, DATA!$C$2:$C${max_data_row}, {cond_y}, DATA!$D$2:$D${max_data_row}, {cond_t}, DATA!$E$2:$E${max_data_row}, {cond_d}), 0.0)'
    appliquer_carte_kpi(ws, 7, 5, "STRESS MÉCANIQUE MOYEN", formule_stress, "0.0")

    # KPI 3 : ALERTES ZONE ROUGE (Interroge et compte la valeur "Zone Rouge (Danger)" en colonne V de DATA)
    ws.merge_cells(start_row=7, start_column=8, end_row=7, end_column=9)
    ws.merge_cells(start_row=8, start_column=8, end_row=8, end_column=9)
    formule_alertes = f'=COUNTIFS(DATA!$V$2:$V${max_data_row}, "Zone Rouge (Danger)", DATA!$C$2:$C${max_data_row}, {cond_y}, DATA!$D$2:$D${max_data_row}, {cond_t}, DATA!$E$2:$E${max_data_row}, {cond_d})'
    appliquer_carte_kpi(ws, 7, 8, "ALERTES ZONE ROUGE", formule_alertes, "0")

def integrer_graphiques_charge(wb, ws_charge, max_data_row):
    """Construit les graphiques en les connectant aux bonnes sources"""
    ws_pivot = wb["Data_Pivot"]
    max_charge_row = 11 + max_data_row - 1
    
    # 1. Le Graphique Circulaire (Interroge Data_Pivot)
    chart_pie = configurer_graphique_pro(PieChart())
    chart_pie.title = "Proportion par Type d'Entraînement"
    chart_pie.height, chart_pie.width = 15, 23
    chart_pie.add_data(Reference(ws_pivot, min_col=6, min_row=1, max_row=4), titles_from_data=True)
    chart_pie.set_categories(Reference(ws_pivot, min_col=5, min_row=2, max_row=4))
    chart_pie.dataLabels = DataLabelList(showPercent=True, showVal=False)
    ws_charge.add_chart(chart_pie, "K11")

    # 2. Le Graphique Linéaire ACWR (Interroge la colonne 7 / G du tableau local)
    chart_line = configurer_graphique_pro(LineChart())
    chart_line.title = "Suivi Chronologique du Ratio ACWR"
    chart_line.height, chart_line.width = 15, 23
    chart_line.y_axis.title = "Ratio de Charge"
    
    chart_line.add_data(Reference(ws_charge, min_col=7, min_row=11, max_row=max_charge_row), titles_from_data=True)
    chart_line.set_categories(Reference(ws_charge, min_col=2, min_row=12, max_row=max_charge_row))
    if chart_line.series: 
        chart_line.series[0].graphicalProperties.line.solidFill = "C0392B"
        chart_line.series[0].smooth = True 
    ws_charge.add_chart(chart_line, "K39")

def construire_page_charge(wb, max_data_row, len_y, len_t, len_d):
    """Construit la page complète de Suivi de Charge en stricte conformité avec ton tableau à 7 colonnes"""
    ws_charge = wb.create_sheet(title="Suivi_Charge")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for row in ws_charge.iter_rows(min_row=1, max_row=250, min_col=1, max_col=22):
        for cell in row: cell.fill = fill_white
    ws_charge.sheet_view.showGridLines = False

    # Configuration des listes déroulantes
    ajouter_filtre_deroulant(ws_charge, 2, 2, "Année", f"=RESSOURCES!$A$1:$A${len_y}", "Tous")
    ajouter_filtre_deroulant(ws_charge, 2, 5, "Trimestre", f"=RESSOURCES!$B$1:$B${len_t}", "Tous")
    ajouter_filtre_deroulant(ws_charge, 2, 8, "Distance Cible", f"=RESSOURCES!$C$1:$C${len_d}", "Tous")

    # Titres de la page
    appliquer_titre_principal(ws_charge.cell(row=4, column=2), "Indicateurs de Charge de Travail (Modèle ACWR)")
    ws_charge.merge_cells(start_row=5, start_column=2, end_row=5, end_column=10)
    appliquer_texte_intro(ws_charge.cell(row=5, column=2), "Analyse de la fatigue aiguë et du risque mécanique.")

    # En-têtes du tableau (Strictement 7 colonnes : B à H comme dans ton code historique)
    start_row_table = 11
    headers = ["Date", "Année", "Trimestre", "Distance (km)", "Stress", "Ratio ACWR", "Statut Alerte"]
    for col_idx, text in enumerate(headers, start=2): 
        ws_charge.cell(row=start_row_table, column=col_idx, value=text)
        
    # Remplissage du tableau local (Formules de liaison vers l'onglet DATA)
    current_row = start_row_table + 1
    for i in range(2, max_data_row + 1):
        cond = f'AND(OR(Suivi_Charge!$B$2="Tous", DATA!C{i}&""=Suivi_Charge!$B$2&""), OR(Suivi_Charge!$E$2="Tous", DATA!D{i}&""=Suivi_Charge!$E$2&""), OR(Suivi_Charge!$H$2="Tous", DATA!E{i}&""=Suivi_Charge!$H$2&""))'
        
        ws_charge.cell(row=current_row, column=2, value=f'=IF({cond}, DATA!B{i}, "")') 
        ws_charge.cell(row=current_row, column=3, value=f'=IF({cond}, DATA!C{i}, "")') 
        ws_charge.cell(row=current_row, column=4, value=f'=IF({cond}, DATA!D{i}, "")') 
        ws_charge.cell(row=current_row, column=5, value=f'=IF({cond}, DATA!F{i}, "")') 
        ws_charge.cell(row=current_row, column=6, value=f'=IF({cond}, DATA!J{i}, "")') 
        ws_charge.cell(row=current_row, column=7, value=f'=IF({cond}, DATA!U{i}, "")') # Ratio ACWR en G
        ws_charge.cell(row=current_row, column=8, value=f'=IF({cond}, DATA!V{i}, "")') # Statut Alerte en H
        
        ws_charge.cell(row=current_row, column=2).number_format = 'YYYY-MM-DD'
        ws_charge.cell(row=current_row, column=5).number_format = '0.0'
        ws_charge.cell(row=current_row, column=6).number_format = '0.0'
        ws_charge.cell(row=current_row, column=7).number_format = '0.00'
        current_row += 1

    max_charge_row = current_row - 1
    styliser_tableau_donnees(ws_charge, start_row_table, max_charge_row, start_col=2, end_col=8)
    
    # Mises en forme conditionnelles appliquées sur la colonne H (colonne 8 / Statut Alerte)
    red_font = Font(color="9C0006", bold=True)
    red_fill = PatternFill(bgColor="FFC7CE", fill_type="solid")
    green_font = Font(color="006100", bold=True)
    green_fill = PatternFill(bgColor="C6EFCE", fill_type="solid")
    
    ws_charge.conditional_formatting.add(f"H12:H{max_charge_row}", CellIsRule(operator='equal', formula=['"Zone Rouge (Danger)"'], font=red_font, fill=red_fill))
    ws_charge.conditional_formatting.add(f"H12:H{max_charge_row}", CellIsRule(operator='equal', formula=['"Zone Optimale"'], font=green_font, fill=green_fill))
    
    # Déclaration de la structure de table Excel standard (B11:H{max})
    tab_charge = Table(displayName="TableauSuiviCharge", ref=f"B11:H{max_charge_row}")
    tab_charge.tableStyleInfo = TableStyleInfo(name="TableStyleMedium14", showRowStripes=True)
    ws_charge.add_table(tab_charge)

    # Déclenchement des KPIs et des graphiques locaux
    injecter_kpis_charge(ws_charge, max_data_row)
    integrer_graphiques_charge(wb, ws_charge, max_data_row)

    # Réglage des dimensions géométriques des cellules
    ws_charge.column_dimensions['A'].width = 3
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']: 
        ws_charge.column_dimensions[col].width = 16