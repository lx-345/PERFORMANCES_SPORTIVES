from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

# Importation des outils de design depuis notre module local
from .styles import (
    ajouter_filtre_deroulant, 
    appliquer_titre_principal, 
    appliquer_texte_intro, 
    appliquer_carte_kpi, 
    styliser_tableau_donnees
)

def construire_page_bilan(wb, max_data_row, len_y, len_s, len_n):
    """Construit l'onglet principal d'analyse du rendement hebdomadaire."""
    ws_bilan = wb.create_sheet(title="Bilan_Hebdo", index=0)
    
    # Fond blanc pour masquer la grille Excel
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for row in ws_bilan.iter_rows(min_row=1, max_row=200, min_col=1, max_col=15):
        for cell in row: cell.fill = fill_white
    ws_bilan.sheet_view.showGridLines = False

    # Filtres interactifs
    ajouter_filtre_deroulant(ws_bilan, 2, 2, "Année", f"=RESSOURCES!$A$1:$A${len_y}", "Tous")
    ajouter_filtre_deroulant(ws_bilan, 2, 4, "Semaine", f"=RESSOURCES!$D$1:$D${len_s}", "Tous")
    ajouter_filtre_deroulant(ws_bilan, 2, 6, "Nom Entraînement", f"=RESSOURCES!$E$1:$E${len_n}", "Tous")

    # En-têtes
    appliquer_titre_principal(ws_bilan.cell(row=4, column=2), "Bilan & Rendement de l'entraînement")
    ws_bilan.merge_cells(start_row=5, start_column=2, end_row=5, end_column=10)
    appliquer_texte_intro(ws_bilan.cell(row=5, column=2), "Analyse de la productivité basée sur l'effort, le stress mécanique et l'ACWR.")

    # KPI : Rendement de la dernière séance affichée
    max_bilan_row = 11 + max_data_row - 1
    ws_bilan.merge_cells(start_row=7, start_column=2, end_row=7, end_column=5)
    ws_bilan.merge_cells(start_row=8, start_column=2, end_row=8, end_column=5)
    
    formule_kpi = f'=IFERROR(LOOKUP(2, 1/(G12:G{max_bilan_row}<>""), G12:G{max_bilan_row}), "Aucune donnée")'
    appliquer_carte_kpi(ws_bilan, 7, 2, "RENDEMENT DE LA DERNIÈRE SEMAINE/SÉANCE", formule_kpi)

    # Construction du tableau de données
    start_row_table = 11
    headers = ["Date", "Année", "Trimestre", "Jour", "Nom d'entraînement", "Rendement"]
    for col_idx, text in enumerate(headers, start=2): 
        ws_bilan.cell(row=start_row_table, column=col_idx, value=text)
        
    current_row = start_row_table + 1
    for i in range(2, max_data_row + 1):
        # La logique de filtrage Excel pointant vers l'onglet caché DATA
        cond = f'AND(OR(Bilan_Hebdo!$B$2="Tous", DATA!C{i}&""=Bilan_Hebdo!$B$2&""), OR(Bilan_Hebdo!$D$2="Tous", DATA!M{i}&""=Bilan_Hebdo!$D$2&""), OR(Bilan_Hebdo!$F$2="Tous", DATA!A{i}&""=Bilan_Hebdo!$F$2&""))'
        
        ws_bilan.cell(row=current_row, column=2, value=f'=IF({cond}, DATA!B{i}, "")') 
        ws_bilan.cell(row=current_row, column=3, value=f'=IF({cond}, DATA!C{i}, "")') 
        ws_bilan.cell(row=current_row, column=4, value=f'=IF({cond}, DATA!D{i}, "")') 
        ws_bilan.cell(row=current_row, column=5, value=f'=IF({cond}, DATA!L{i}, "")') 
        ws_bilan.cell(row=current_row, column=6, value=f'=IF({cond}, DATA!A{i}, "")') 
        ws_bilan.cell(row=current_row, column=7, value=f'=IF({cond}, DATA!W{i}, "")') 
        
        ws_bilan.cell(row=current_row, column=2).number_format = 'YYYY-MM-DD'
        current_row += 1

    styliser_tableau_donnees(ws_bilan, start_row_table, current_row - 1, start_col=2, end_col=7)
    
    # Formatage conditionnel (Couleurs Vert / Rouge pour le rendement)
    red_font = Font(color="9C0006", bold=True)
    red_fill = PatternFill(bgColor="FFC7CE", fill_type="solid")
    green_font = Font(color="006100", bold=True)
    green_fill = PatternFill(bgColor="C6EFCE", fill_type="solid")
    
    ws_bilan.conditional_formatting.add("B8:E8", CellIsRule(operator='equal', formula=['"Semaine non productive"'], font=red_font, fill=red_fill))
    ws_bilan.conditional_formatting.add("B8:E8", CellIsRule(operator='equal', formula=['"Semaine productive"'], font=green_font, fill=green_fill))
    ws_bilan.conditional_formatting.add(f"G12:G{max_bilan_row}", CellIsRule(operator='equal', formula=['"Semaine non productive"'], font=red_font, fill=red_fill))
    ws_bilan.conditional_formatting.add(f"G12:G{max_bilan_row}", CellIsRule(operator='equal', formula=['"Semaine productive"'], font=green_font, fill=green_fill))
    
    # Création du tableau officiel Excel
    tab = Table(displayName="TableauBilan", ref=f"B11:G{max_bilan_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium14", showRowStripes=True)
    ws_bilan.add_table(tab)
    
    # Largeur des colonnes
    for col in ['B', 'C', 'D', 'E']: 
        ws_bilan.column_dimensions[col].width = 15
    ws_bilan.column_dimensions['F'].width = 30
    ws_bilan.column_dimensions['G'].width = 25