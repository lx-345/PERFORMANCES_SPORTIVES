import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

def construire_onglets_back(wb, df_export, trimestres_uniques, max_data_row):
    """Construit les 3 onglets techniques cachés (DATA, Data_Pivot, RESSOURCES)"""
    
    # --------------------------------------------------------------------------
    # AJOUT DU FILTRE : Exclusion stricte de la période 2024
    # --------------------------------------------------------------------------
    mask_hors_2024 = pd.to_numeric(df_export['annee'], errors='coerce') != 2024
    df_export = df_export[mask_hors_2024].copy()
    
    # On recalcule immédiatement les dimensions pour que tout le reste du classeur s'ajuste
    max_data_row = len(df_export) + 1
    trimestres_uniques = sorted(df_export['trimestre'].dropna().unique().astype(str).tolist())
    # --------------------------------------------------------------------------

    # Forcer le type Datetime pour les calculs de charge chronologique
    df_export['start_date_local'] = pd.to_datetime(df_export['start_date_local'], errors='coerce')
    
    # 1. Onglet DATA (Moteur de calcul)
    ws_data = wb.create_sheet(title="DATA")
    for r in dataframe_to_rows(df_export, index=False, header=True): 
        ws_data.append(r)
    
    headers_calc = ["Y_Intense", "Y_Modere", "Y_Faible", "Y_Temps", "Y_Effort", "Charge_Aigue", "Charge_Chronique", "Ratio_ACWR", "Statut Alerte", "Rendement"]
    for idx, val in enumerate(headers_calc, start=14): 
        ws_data.cell(row=1, column=idx, value=val)

    for i in range(2, max_data_row + 1):
        cond = 'OR(Analyse!$B$2="Tous", $C{0}&""=Analyse!$B$2&""), OR(Analyse!$E$2="Tous", $D{0}&""=Analyse!$E$2&""), OR(Analyse!$H$2="Tous", $E{0}&""=Analyse!$H$2&"")'.format(i)
        
        ws_data.cell(row=i, column=14, value=f'=IF(AND($K{i}="Intense", {cond}), $J{i}, NA())')
        ws_data.cell(row=i, column=15, value=f'=IF(AND($K{i}="Modéré", {cond}), $J{i}, NA())')
        ws_data.cell(row=i, column=16, value=f'=IF(AND($K{i}="Faible", {cond}), $J{i}, NA())')
        ws_data.cell(row=i, column=17, value=f'=IF(AND({cond}), $G{i}, NA())')
        ws_data.cell(row=i, column=18, value=f'=IF(AND({cond}), $I{i}, NA())')
        
        ws_data.cell(row=i, column=19, value=f'=SUMIFS(J:J, B:B, ">="&B{i}-6, B:B, "<="&B{i})')
        ws_data.cell(row=i, column=20, value=f'=SUMIFS(J:J, B:B, ">="&B{i}-41, B:B, "<="&B{i}) / 6')
        ws_data.cell(row=i, column=21, value=f'=IFERROR(S{i} / T{i}, 1)')
        ws_data.cell(row=i, column=22, value=f'=IF(U{i}>1.5, "Zone Rouge (Danger)", IF(U{i}>=0.8, "Zone Optimale", "Sous-charge"))')
        ws_data.cell(row=i, column=23, value=f'=IF(U{i}>1.5, "Semaine non productive", IF(U{i}>=0.8, "Semaine productive", "Semaine non productive"))')
        
        ws_data.cell(row=i, column=2).number_format = 'YYYY-MM-DD'

    tab = Table(displayName="RawData", ref=f"A1:W{max_data_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium14", showRowStripes=True)
    ws_data.add_table(tab)

    # 2. Onglet Data_Pivot (Agrégations pour graphiques)
    ws_pivot = wb.create_sheet(title="Data_Pivot")
    ws_pivot["A1"], ws_pivot["B1"], ws_pivot["C1"] = "Trimestre", "Volume Total", "VMA Moyenne"
    
    max_dash_row = 44 + max_data_row - 1
    for i, trim in enumerate(trimestres_uniques, start=2):
        ws_pivot.cell(row=i, column=1, value=trim)
        ws_pivot.cell(row=i, column=2, value=f'=SUMIFS(Analyse!$F$45:$F${max_dash_row}, Analyse!$D$45:$D${max_dash_row}, A{i})')
        ws_pivot.cell(row=i, column=3, value=f'=IFERROR(AVERAGEIFS(Analyse!$H$45:$H${max_dash_row}, Analyse!$D$45:$D${max_dash_row}, A{i}), NA())')
    
    ws_pivot["E1"], ws_pivot["F1"] = "Type", "Nombre"
    ws_pivot["E2"], ws_pivot["E3"], ws_pivot["E4"] = "Intense", "Modéré", "Faible"
    
    cond_y_pie = 'IF(Suivi_Charge!$B$2="Tous", "*", Suivi_Charge!$B$2)'
    cond_t_pie = 'IF(Suivi_Charge!$E$2="Tous", "*", Suivi_Charge!$E$2)'
    cond_d_pie = 'IF(Suivi_Charge!$H$2="Tous", "*", Suivi_Charge!$H$2)'

    ws_pivot["F2"] = f'=COUNTIFS(DATA!$K$2:$K${max_data_row}, "Intense", DATA!$C$2:$C${max_data_row}, {cond_y_pie}, DATA!$D$2:$D${max_data_row}, {cond_t_pie}, DATA!$E$2:$E${max_data_row}, {cond_d_pie})'
    ws_pivot["F3"] = f'=COUNTIFS(DATA!$K$2:$K${max_data_row}, "Modéré", DATA!$C$2:$C${max_data_row}, {cond_y_pie}, DATA!$D$2:$D${max_data_row}, {cond_t_pie}, DATA!$E$2:$E${max_data_row}, {cond_d_pie})'
    ws_pivot["F4"] = f'=COUNTIFS(DATA!$K$2:$K${max_data_row}, "Faible", DATA!$C$2:$C${max_data_row}, {cond_y_pie}, DATA!$D$2:$D${max_data_row}, {cond_t_pie}, DATA!$E$2:$E${max_data_row}, {cond_d_pie})'

    # 3. Onglet RESSOURCES (Listes déroulantes)
    ws_res = wb.create_sheet(title="RESSOURCES")
    annees = ["Tous"] + sorted(df_export['annee'].dropna().unique().astype(str).tolist())
    trims = ["Tous"] + sorted(df_export['trimestre'].dropna().unique().astype(str).tolist())
    dists = ["Tous"] + sorted(df_export['distance_cible'].dropna().unique().tolist())
    
    df_export['semaine'] = pd.to_numeric(df_export['semaine'], errors='coerce').fillna(0).astype(int)
    semaines = ["Tous"] + sorted(df_export['semaine'].unique().tolist())
    if 0 in semaines: semaines.remove(0)
    semaines = [str(s) for s in semaines]
    
    noms = ["Tous"] + sorted(df_export['name'].dropna().unique().tolist())
    
    for row, val in enumerate(annees, start=1): ws_res.cell(row=row, column=1, value=val)
    for row, val in enumerate(trims, start=1): ws_res.cell(row=row, column=2, value=val)
    for row, val in enumerate(dists, start=1): ws_res.cell(row=row, column=3, value=val)
    for row, val in enumerate(semaines, start=1): ws_res.cell(row=row, column=4, value=val)
    for row, val in enumerate(noms, start=1): ws_res.cell(row=row, column=5, value=val)
        
    return len(annees), len(trims), len(dists), len(semaines), len(noms), len(trimestres_uniques) + 1