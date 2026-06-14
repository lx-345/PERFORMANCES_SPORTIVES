from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def appliquer_titre_principal(cell, texte):
    cell.value = texte
    cell.font = Font(name='Segoe UI', size=18, bold=True, color="E16B1A")
    cell.alignment = Alignment(horizontal='left', vertical='center')

def appliquer_texte_intro(cell, texte):
    cell.value = texte
    cell.font = Font(name='Segoe UI', size=10, italic=True, color="595959")
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

def appliquer_style_filtre(cell_titre, cell_valeur, titre):
    cell_titre.value = titre
    cell_titre.font = Font(name='Segoe UI', size=9, bold=True, color="595959")
    cell_titre.alignment = Alignment(horizontal='center', vertical='center')
    
    cell_valeur.font = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
    cell_valeur.fill = PatternFill(start_color="E16B1A", end_color="E16B1A", fill_type="solid")
    cell_valeur.alignment = Alignment(horizontal='center', vertical='center')
    
    border_thin = Border(left=Side(style='thin', color="BFBFBF"), right=Side(style='thin', color="BFBFBF"), 
                         top=Side(style='thin', color="BFBFBF"), bottom=Side(style='thin', color="BFBFBF"))
    cell_valeur.border = border_thin

def ajouter_filtre_deroulant(ws, row, col, titre, formule_ressource, valeur_defaut):
    appliquer_style_filtre(ws.cell(row=row-1, column=col), ws.cell(row=row, column=col), titre)
    cell_valeur = ws.cell(row=row, column=col)
    cell_valeur.value = valeur_defaut
    dv = DataValidation(type="list", formula1=formule_ressource, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_valeur)

def appliquer_carte_kpi(ws, row, col, titre, formule, format_nb=None):
    cell_titre = ws.cell(row=row, column=col)
    cell_valeur = ws.cell(row=row+1, column=col)
    
    cell_titre.value = titre
    cell_titre.font = Font(name='Segoe UI', size=10, bold=True, color="FFFFFF")
    cell_titre.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    cell_titre.alignment = Alignment(horizontal='center', vertical='center')
    
    cell_valeur.value = formule
    cell_valeur.font = Font(name='Segoe UI', size=16, bold=True, color="E16B1A")
    cell_valeur.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    cell_valeur.alignment = Alignment(horizontal='center', vertical='center')
    
    if format_nb: cell_valeur.number_format = format_nb
        
    border_thin = Border(left=Side(style='thin', color="D9D9D9"), right=Side(style='thin', color="D9D9D9"), 
                         top=Side(style='thin', color="D9D9D9"), bottom=Side(style='thin', color="D9D9D9"))
    for r in range(row, row+2):
        for c in range(col, col+2): ws.cell(row=r, column=c).border = border_thin

def styliser_tableau_donnees(ws, start_row, max_row, start_col=2, end_col=9):
    header_font = Font(name='Segoe UI', size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
    data_font = Font(name='Segoe UI', size=10, color="333333")
    border_light = Border(bottom=Side(style='thin', color="E0E0E0"))
    
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r in range(start_row + 1, max_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = data_font
            cell.border = border_light
            if col in [2, 3, 4, 5, 12]: cell.alignment = Alignment(horizontal='center', vertical='center')
            else: cell.alignment = Alignment(horizontal='right', vertical='center')

def configurer_graphique_pro(chart):
    chart.style = 14
    if chart.legend:
        chart.legend.position = 'b'
    return chart